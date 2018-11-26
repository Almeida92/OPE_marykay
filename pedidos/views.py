from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os.path
from pedidos.models import Pedido, PedidoProduto
from consultoras.models import Consultora
from produtos.models import Produto
from django.db import models, IntegrityError
from django.shortcuts import redirect
import datetime

def consultar_pedidos(request):

  lista_pedidos = PedidoProduto.objects.all()
  consultora = Consultora.objects.get(usuario_id = request.user.id)
  lista = []
  for pedido in lista_pedidos:
    if (pedido.pedido.consultora.id == consultora.id):
      lista.append(pedido)

  return render(request, 'pedidos/consultar.html', {'lista_pedidos' : lista})

def importar_pedidos(request):
  if request.method == 'POST':
      myfile = request.FILES['file']
      try:
        consultora = Consultora.objects.get(usuario = request.user.id)
        salvar_arquivo(myfile, consultora)
      except:
        print("Erro")
  else:
    pass
    
  return render(request, 'pedidos/importar.html')

def salvar_arquivo(arquivo, consultora):
  folder='tmp/'
  fs = FileSystemStorage(location=folder)
  filename = fs.save(arquivo.name, arquivo)
  file_url = fs.url(filename)
  ler_arquivo(file_url, consultora)

def ler_arquivo(nome_arquivo, consultora):
  filename = 'tmp/' + nome_arquivo
  wb = load_workbook(filename)

  sheet_ranges = wb.worksheets[0]
  limpar_planilha(sheet_ranges, filename, consultora)

def limpar_planilha(planilha, filename, consultora):
  linhas = []
  for index, linha in enumerate(planilha):
    if index > 0:
      linhas.append(linha)

  remover_excel(filename)

  ler_colunas(linhas, consultora)

def ler_colunas(linhas, consultora):
  produtos = []
  for linha in linhas:
    produto = {}
    produto['codigo'] = linha[0].value
    produto['descricao'] = linha[1].value
    produto['preco'] = linha[2].value
    produto['pontos'] = linha[3].value
    produto['quantidade'] = linha[4].value
    produto['total'] = linha[5].value
    produtos.append(produto)

  salvar_pedidos(produtos, consultora)

def salvar_pedidos(produtos, consultora):
  novo_pedido = Pedido(
    consultora = consultora,
    data_pedido = datetime.date.today()
  )

  try:
    novo_pedido.save()
  except:
    print("Erro ao salvar Pedido")
  
  for produto in produtos:
    try:
        produto_cadastrado = Produto.objects.get(codigo = produto['codigo'])
    except Produto.DoesNotExist:
        produto_cadastrado = None

    if (produto_cadastrado):
      novo_pedido_produto = PedidoProduto(
        quantidade = produto['quantidade'],
        produto = produto_cadastrado,
        pedido = novo_pedido)

      novo_pedido_produto.save()

def remover_excel(filename):
  os.remove(filename)

def pedido_detalhe(request , pedido):
  pedido = Pedido.objects.get(id=pedido)
  pedido_produto = PedidoProduto.objects.filter(pedido_id=pedido)
  
  return render(request, 'pedidos/detalhe.html', {
    'pedido': pedido,
    'pedido_produto': pedido_produto
  })